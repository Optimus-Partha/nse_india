from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time, datetime, os
import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service


def stock_excel_update():
    user = os.path.expanduser("~").replace("C:\\Users\\", "")

    todays_excel = pd.DataFrame(
        {"Calls OI": [], "Calls CHNG IN OI": [], "Calls VOLUME": [], "Calls IV": [], "Calls LTP": [],
         "Calls CHNG": [], "Calls BID QTY": [], "Calls BID PRICE": [], "Calls ASK PRICE": [],
         "Calls ASK QTY": [], "Puts STRIKE PRICE": [], "Puts BID QTY": [], "Puts BID PRICE": [],
         "Puts ASK PRICE": [], "Puts ASK QTY": [], "Puts CHNG": [], "Puts LTP": [], "Puts IV": [],
         "Puts VOLUME": [], "Puts CHNG IN OI": [], "Puts OI": [], "Nifty": [], "Time": [], "Date": []})

    todays_excel_path = "StockMarket_" + str(datetime.date.today()).replace('-', '_') + '.xlsx'

    if not os.path.exists('.\\' + todays_excel_path):
        todays_excel.to_excel(todays_excel_path, index=False)
        print('New excel created for today.')

    def every_downloads_chrome(driver):
        if not driver.current_url.startswith("chrome://downloads"):
            driver.get("chrome://downloads/")
        return driver.execute_script("""
            return document.querySelector('downloads-manager')
            .shadowRoot.querySelector('#downloadsList')
            .items.filter(e => e.state === 'COMPLETE')
            .map(e => e.filePath || e.file_path || e.fileUrl || e.file_url);
            """)

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    # options.add_argument('--headless')
    # driver = webdriver.Chrome(ChromeDriverManager().install())
    s = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s)
    # driver = webdriver.Chrome('chromedriver.exe')
    driver.get("https://www.nseindia.com/option-chain")
    WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.ID, "equity_underlyingVal")))
    nifty = (driver.find_element(By.XPATH, '//*[@id="equity_underlyingVal"]').text).replace('NIFTY ', '').replace(',',
                                                                                                                  '')
    time_stamp = driver.find_element(By.XPATH, '//*[@id="equity_timeStamp"]').text
    WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.XPATH, '//*[@id="optionChainTable-indices"]/tbody/tr[1]/td[2]')))
    driver.find_element(By.XPATH, '//*[@id="downloadOCTable"]').click()
    paths = WebDriverWait(driver, 60, 1).until(every_downloads_chrome)
    driver.quit()
    data_dump = pd.read_csv(paths[0], skiprows=1, low_memory=False)
    data_dump['NIFTY'] = nifty
    data_dump['Time'] = time_stamp.split(" ")[3]
    data_dump['Date'] = time_stamp.split(" ")[2]
    print(nifty, time_stamp)
    data_dump.drop(['Unnamed: 0', 'Unnamed: 22'], axis='columns', inplace=True)

    data_dump['STRIKE PRICE'] = data_dump['STRIKE PRICE'].str.replace(',', '')
    data_dump['STRIKE PRICE'] = pd.to_numeric(data_dump['STRIKE PRICE'], errors='coerce')
    result_index = data_dump['STRIKE PRICE'].sub(float(nifty)).abs().idxmin()
    data_dump = data_dump.loc[result_index - 4:result_index + 4]

    running_dump = pd.read_excel(todays_excel_path)

    if running_dump.empty == True or max(running_dump['Time_Stamp']) < max(data_dump['Time_Stamp']):
        book = load_workbook(todays_excel_path)
        writer = pd.ExcelWriter(todays_excel_path, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        data_dump.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)
        writer.save()
        print("New Data Added")
    os.remove(paths[0])


def update_excel():
    try:
        stock_excel_update()
    except:
        pass


if __name__ == '__main__':
    # stock_excel_update()
    update_excel()
